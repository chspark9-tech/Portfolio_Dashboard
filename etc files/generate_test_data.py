import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import datetime
import random

def generate_test_xlsx():
    wb = openpyxl.Workbook()
    
    # 색상 및 스타일 설정
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    def style_header(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    # 1. 0.인덱스
    ws_idx = wb.active
    ws_idx.title = "0.인덱스"
    ws_idx.append(["계좌내역", "드롭박스(계좌이름)"])
    accounts = ["연금계좌", "ISA", "삼성미국", "IRP"]
    for i in range(1, 11):
        name = accounts[i-1] if i <= len(accounts) else ""
        ws_idx.append([f"계좌내역{i}", name])
    style_header(ws_idx)

    # 2. 종목현황
    ws_status = wb.create_sheet("종목현황")
    headers_status = ["국가", "계좌종류", "분류", "종목코드", "종목명", "수량", "평단가(원화)", "평단가(달러)"]
    ws_status.append(headers_status)
    
    stocks = [
        ["한국", "연금계좌", "지수", "005930", "삼성전자", 100, 72000, ""],
        ["한국", "연금계좌", "배당", "055550", "신한지주", 200, 45000, ""],
        ["한국", "ISA", "지수", "379800", "KODEX 미국S&P500", 500, 15000, ""],
        ["미국", "삼성미국", "기술주", "AAPL", "Apple", 20, "", 190.5],
        ["미국", "삼성미국", "기술주", "MSFT", "Microsoft", 10, "", 420.0],
        ["미국", "삼성미국", "배당", "SCHD", "Schwab US Dividend Equity", 100, "", 78.0],
    ]
    for s in stocks:
        ws_status.append(s)
    style_header(ws_status)

    # 3. 배당내역
    ws_div = wb.create_sheet("배당내역")
    ws_div.append(["일자", "계좌", "종목코드", "종목명", "원화배당금", "외화배당금"])
    
    start_date = datetime.date(2023, 1, 1)
    end_date = datetime.date(2026, 4, 1)
    
    current = start_date
    while current <= end_date:
        full_date_str = current.strftime("%Y-%m-%d")
        # 분기별 배당 시뮬레이션
        if current.month in [3, 6, 9, 12]:
            # 한국 배당
            ws_div.append([full_date_str, "연금계좌", "005930", "삼성전자", 36100, 0])
            ws_div.append([full_date_str, "연금계좌", "055550", "신한지주", 50000, 0])
        
        # 매달 미국 배당 시뮬레이션 (일부 종목)
        if current.month % 3 == 0:
            ws_div.append([full_date_str, "삼성미국", "SCHD", "Schwab US Dividend Equity", 0, 65.5])
            ws_div.append([full_date_str, "삼성미국", "AAPL", "Apple", 0, 4.8])
            
        # 다음달로 이동
        if current.month == 12:
            current = datetime.date(current.year + 1, 1, 1)
        else:
            current = datetime.date(current.year, current.month + 1, 1)
    style_header(ws_div)

    # 4. 입금내역 (2023~2026 데이터 생성)
    ws_dep = wb.create_sheet("입금내역")
    ws_dep.append(["날짜", "계좌", "입금액", "비고"])
    
    current = start_date
    while current <= end_date:
        full_date_str = current.strftime("%Y-%m-%d")
        for acc in accounts:
            ws_dep.append([full_date_str, acc, 500000, "정기입금"])
            if random.random() > 0.8:
                ws_dep.append([full_date_str, acc, 2000000, "보너스입금"])
        
        # 다음달로 이동
        if current.month == 12:
            current = datetime.date(current.year + 1, 1, 1)
        else:
            current = datetime.date(current.year, current.month + 1, 1)
    style_header(ws_dep)

    # 5. 계좌내역 (1~4번 계좌만 데이터 채움)
    for i in range(1, 11):
        ws_acc = wb.create_sheet(f"계좌내역{i}")
        ws_acc.append(["날짜", "월말평가액"])
        
        if i <= len(accounts):
            acc_name = accounts[i-1]
            curr_balance = 10000000 * i
            
            curr_date = start_date
            while curr_date <= end_date:
                # 계좌내역은 YY/MM 형식 사용
                yy_mm_str = curr_date.strftime("%y/%m")
                curr_balance = (curr_balance + 500000) * (1 + random.uniform(-0.02, 0.05))
                ws_acc.append([yy_mm_str, curr_balance])
                
                if curr_date.month == 12:
                    curr_date = datetime.date(curr_date.year + 1, 1, 1)
                else:
                    curr_date = datetime.date(curr_date.year, curr_date.month + 1, 1)
        style_header(ws_acc)

    file_path = r"c:\AI_Project\Portpolio_dashboard\계좌 현황 기록시트_test.xlsx"
    wb.save(file_path)
    print(f"Test data generated at: {file_path}")

if __name__ == "__main__":
    generate_test_xlsx()
