import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_portfolio_template():
    wb = openpyxl.Workbook()

    # ── 공통 스타일 ──
    header_fill  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    sample_fill  = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    guide_fill   = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    sample_font  = Font(name="Arial", color="2C5282", size=10)
    guide_font   = Font(name="Arial", color="808080", italic=True, size=9)
    body_font    = Font(name="Arial", size=10)
    thin_side    = Side(style="thin", color="CCCCCC")
    thin_border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def style_header_row(ws, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[1].height = 20

    def style_sample_rows(ws, start_row, end_row, col_count):
        for r in range(start_row, end_row + 1):
            for c in range(1, col_count + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = sample_fill
                cell.font = sample_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

    def style_guide_row(ws, row, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = guide_fill
            cell.font = guide_font
            cell.border = thin_border

    def set_col_widths(ws, widths):
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ════════════════════════════════════════════
    # 1. 0.인덱스
    # ════════════════════════════════════════════
    ws_idx = wb.active
    ws_idx.title = "0.인덱스"

    # ── 헤더: A~G열 ──
    # app.js exportWorkbook: D1='계좌합1 자산군', E1='비중(%)', F1='계좌합2 자산군', G1='비중(%)'
    # app.js parseWorkbook:  row[3]=gd1 자산군, row[4]=gd1 비중, row[5]=gd2 자산군, row[6]=gd2 비중
    ws_idx.append(["계좌내역", "드롭박스(계좌이름)", "",
                   "계좌합1 자산군", "비중(%)", "계좌합2 자산군", "비중(%)"])

    # A,B 헤더 스타일
    for c in [1, 2]:
        cell = ws_idx.cell(row=1, column=c)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # C열 빈 구분자
    ws_idx.cell(1, 3).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    # D~G 목표비중 헤더 (두 그룹 색상 구분)
    goal_fill1 = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")  # 진초록 - 계좌합1
    goal_fill2 = PatternFill(start_color="4A235A", end_color="4A235A", fill_type="solid")  # 진보라 - 계좌합2
    for c, fill in [(4, goal_fill1), (5, goal_fill1), (6, goal_fill2), (7, goal_fill2)]:
        cell = ws_idx.cell(row=1, column=c)
        cell.fill = fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_idx.row_dimensions[1].height = 20

    # ── 계좌내역 데이터 (A,B열, 행2~11) ──
    accounts = ["연금저축", "ISA계좌", "미국직투", "IRP", "기타계좌", "", "", "", "", ""]
    for i in range(1, 11):
        ws_idx.append([f"계좌내역{i}", accounts[i-1]])
        for c in range(1, 3):
            cell = ws_idx.cell(row=i+1, column=c)
            cell.font = body_font; cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # ── 목표비중 샘플 데이터 (D,E,F,G열, 행2~) ──
    # app.js: idx[r][3]=자산군1, idx[r][4]=비중1, idx[r][5]=자산군2, idx[r][6]=비중2
    goal_sample_fill1 = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    goal_sample_fill2 = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
    goal_sample_font1 = Font(name="Arial", color="1B5E20", size=10)
    goal_sample_font2 = Font(name="Arial", color="4A235A", size=10)

    goal_data = [
        # (자산군1,    비중1,  자산군2,    비중2)
        ("국내주식",    30,    "국내주식",   20),
        ("미국주식",    40,    "미국주식",   50),
        ("채권/현금",   20,    "배당ETF",    20),
        ("대안자산",    10,    "채권/현금",  10),
    ]

    for row_i, (a1, w1, a2, w2) in enumerate(goal_data, start=2):
        ws_idx.cell(row_i, 4).value = a1
        ws_idx.cell(row_i, 5).value = w1
        ws_idx.cell(row_i, 6).value = a2
        ws_idx.cell(row_i, 7).value = w2
        for c in [4, 5]:
            cell = ws_idx.cell(row_i, c)
            cell.fill = goal_sample_fill1; cell.font = goal_sample_font1
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if c == 5 else "left", vertical="center")
        for c in [6, 7]:
            cell = ws_idx.cell(row_i, c)
            cell.fill = goal_sample_fill2; cell.font = goal_sample_font2
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if c == 7 else "left", vertical="center")

    # C열 구분자 스타일 (행 2~11)
    sep_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    for r in range(2, 12):
        ws_idx.cell(r, 3).fill = sep_fill

    set_col_widths(ws_idx, [18, 20, 3, 18, 10, 18, 10])

    # ════════════════════════════════════════════
    # 2. 종목현황
    # app.js parseWorkbook: isNew 형식 (startRow=1)
    # 컬럼 순서: 국가(0) 계좌종류(1) 분류(2) 종목코드(3) 종목명(4)
    #            수량(5) 평단가(원화)(6) 평단가(달러)(7)
    # ════════════════════════════════════════════
    ws_st = wb.create_sheet("종목현황")

    headers_st = ["국가", "계좌종류", "분류", "종목코드", "종목명",
                  "수량", "평단가(원화)", "평단가(달러)"]
    ws_st.append(headers_st)
    style_header_row(ws_st, len(headers_st))

    # 가이드 행
    guide_st = ["한국/미국", "0.인덱스의 계좌이름", "지수/배당/기술주 등", "티커or종목코드",
                "종목 이름", "보유 수량", "한국주식 평균단가(원)", "미국주식 평균단가($)"]
    ws_st.append(guide_st)
    style_guide_row(ws_st, 2, len(headers_st))

    # 샘플 데이터
    samples_st = [
        ["한국", "연금저축", "배당",  "005930", "삼성전자", 50,  72000,  ""],
        ["한국", "ISA계좌",  "지수",  "379800", "KODEX 미국S&P500", 100, 15200, ""],
        ["미국", "미국직투", "기술주", "AAPL",  "Apple Inc.",  10,  "", 189.50],
        ["미국", "미국직투", "배당",  "SCHD",  "Schwab US Dividend ETF", 30, "", 78.20],
    ]
    for row in samples_st:
        ws_st.append(row)
    style_sample_rows(ws_st, 3, 2 + len(samples_st), len(headers_st))

    # 숫자 정렬
    for r in range(3, 3 + len(samples_st)):
        for c in [6, 7, 8]:
            ws_st.cell(r, c).alignment = Alignment(horizontal="right", vertical="center")

    set_col_widths(ws_st, [8, 14, 10, 12, 26, 8, 16, 16])

    # ════════════════════════════════════════════
    # 3. 배당내역
    # app.js renderDividendTab: row[0]=일자, row[1]=계좌, row[4]=원화배당금, row[5]=외화배당금
    # ════════════════════════════════════════════
    ws_div = wb.create_sheet("배당내역")

    headers_div = ["일자", "계좌", "종목코드", "종목명", "원화배당금", "외화배당금"]
    ws_div.append(headers_div)
    style_header_row(ws_div, len(headers_div))

    guide_div = ["YYYY-MM-DD", "0.인덱스의 계좌이름", "티커or종목코드", "종목명",
                 "원화 배당금 (원)", "외화 배당금 ($) — 원화이면 0"]
    ws_div.append(guide_div)
    style_guide_row(ws_div, 2, len(headers_div))

    samples_div = [
        ["2024-03-20", "연금저축", "005930", "삼성전자",       361,  0],
        ["2024-06-20", "연금저축", "005930", "삼성전자",       361,  0],
        ["2024-03-28", "미국직투", "AAPL",   "Apple Inc.",      0,   0.24],
        ["2024-06-14", "미국직투", "AAPL",   "Apple Inc.",      0,   0.25],
        ["2024-03-27", "미국직투", "SCHD",   "Schwab US Dividend ETF", 0, 65.80],
        ["2024-06-26", "미국직투", "SCHD",   "Schwab US Dividend ETF", 0, 67.10],
    ]
    for row in samples_div:
        ws_div.append(row)
    style_sample_rows(ws_div, 3, 2 + len(samples_div), len(headers_div))

    for r in range(3, 3 + len(samples_div)):
        for c in [5, 6]:
            ws_div.cell(r, c).alignment = Alignment(horizontal="right", vertical="center")

    set_col_widths(ws_div, [14, 14, 12, 28, 16, 16])

    # ════════════════════════════════════════════
    # 4. 입금내역
    # app.js renderCumulativeTab: row[0]=날짜, row[1]=계좌, row[2]=입금액
    # ════════════════════════════════════════════
    ws_dep = wb.create_sheet("입금내역")

    headers_dep = ["날짜", "계좌", "입금액", "비고"]
    ws_dep.append(headers_dep)
    style_header_row(ws_dep, len(headers_dep))

    guide_dep = ["YYYY-MM-DD", "0.인덱스의 계좌이름", "입금액 (원)", "메모 (선택)"]
    ws_dep.append(guide_dep)
    style_guide_row(ws_dep, 2, len(headers_dep))

    samples_dep = [
        ["2024-01-02", "연금저축", 500000,  "정기납입"],
        ["2024-01-02", "ISA계좌",  500000,  "정기납입"],
        ["2024-01-02", "미국직투", 1000000, "환전 후 입금"],
        ["2024-02-01", "연금저축", 500000,  "정기납입"],
        ["2024-02-01", "미국직투", 500000,  "정기납입"],
        ["2024-03-01", "연금저축", 500000,  "정기납입"],
        ["2024-03-01", "미국직투", 2000000, "보너스 입금"],
    ]
    for row in samples_dep:
        ws_dep.append(row)
    style_sample_rows(ws_dep, 3, 2 + len(samples_dep), len(headers_dep))

    for r in range(3, 3 + len(samples_dep)):
        ws_dep.cell(r, 3).alignment = Alignment(horizontal="right", vertical="center")

    set_col_widths(ws_dep, [14, 14, 14, 20])

    # ════════════════════════════════════════════
    # 5. 계좌내역1~10
    # app.js renderCumulativeTab: row[0]=날짜(YY/MM), row[1]=월말평가액
    # ════════════════════════════════════════════
    acc_samples = {
        "계좌내역1": {  # 연금저축
            "rows": [
                ["24/01", 10500000], ["24/02", 10950000], ["24/03", 11200000],
                ["24/04", 11600000], ["24/05", 12100000], ["24/06", 12450000],
                ["24/07", 12900000], ["24/08", 13200000], ["24/09", 13500000],
                ["24/10", 13900000], ["24/11", 14300000], ["24/12", 14800000],
            ]
        },
        "계좌내역2": {  # ISA계좌
            "rows": [
                ["24/01", 5200000], ["24/02", 5450000], ["24/03", 5600000],
                ["24/04", 5800000], ["24/05", 6100000], ["24/06", 6250000],
            ]
        },
        "계좌내역3": {  # 미국직투
            "rows": [
                ["24/01", 8500000], ["24/02", 9100000], ["24/03", 9600000],
                ["24/04", 9300000], ["24/05", 9800000], ["24/06", 10400000],
            ]
        },
    }

    for i in range(1, 11):
        ws_acc = wb.create_sheet(f"계좌내역{i}")
        ws_acc.append(["날짜", "월말평가액"])
        style_header_row(ws_acc, 2)

        guide_acc = ["YY/MM 형식 (예: 24/01)", "해당 월 마지막 날 기준 평가금액 (원)"]
        ws_acc.append(guide_acc)
        style_guide_row(ws_acc, 2, 2)

        key = f"계좌내역{i}"
        if key in acc_samples:
            for row in acc_samples[key]["rows"]:
                ws_acc.append(row)
            end_row = 2 + len(acc_samples[key]["rows"])
            style_sample_rows(ws_acc, 3, end_row, 2)
            for r in range(3, end_row + 1):
                ws_acc.cell(r, 2).alignment = Alignment(horizontal="right", vertical="center")

        set_col_widths(ws_acc, [22, 22])

    # ── 저장 ──
    out = "/mnt/user-data/outputs/Portfolio_CheckIn_Template.xlsx"
    wb.save(out)
    print(f"저장 완료: {out}")

if __name__ == "__main__":
    create_portfolio_template()
