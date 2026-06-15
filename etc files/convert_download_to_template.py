import pandas as pd
import openpyxl
import os

def convert_to_template_robust():
    # 경로 설정
    base_path = r'C:\AI_Project\Portpolio_dashboard_1.1'
    source_file = os.path.join(base_path, 'download.xlsx')
    template_file = os.path.join(base_path, 'Portfolio_CheckIn_Template.xlsx')
    output_file = os.path.join(base_path, 'Portfolio_Ready.xlsx')

    if not os.path.exists(source_file):
        print(f"Error: {source_file} 파일을 찾을 수 없습니다.")
        return

    if not os.path.exists(template_file):
        print(f"Error: {template_file} 파일을 찾을 수 없습니다.")
        return

    print("1. 템플릿 로드 중...")
    wb_template = openpyxl.load_workbook(template_file)
    
    print("2. 소스 데이터 로드 중 (이 작업은 다소 시간이 걸릴 수 있습니다)...")
    # 소스 데이터의 모든 시트를 데이터프레임 딕셔너리로 읽어옴
    try:
        source_data = pd.read_excel(source_file, sheet_name=None)
    except Exception as e:
        print(f"Error: 소스 파일을 읽는 중 오류 발생: {e}")
        return

    print("3. 헤더 기반 데이터 매칭 및 복사 중...")
    for sheet_name in wb_template.sheetnames:
        if sheet_name in source_data:
            ws_template = wb_template[sheet_name]
            df_source = source_data[sheet_name]
            
            # 템플릿의 헤더(1행) 읽기
            template_headers = [cell.value for cell in ws_template[1] if cell.value is not None]
            if not template_headers:
                print(f"   ! '{sheet_name}' 시트: 템플릿 헤더를 찾을 수 없어 건너뜁니다.")
                continue

            print(f"   ▶ 시트 매칭됨: '{sheet_name}'")
            
            # 기존 데이터 삭제 (2행부터 끝까지)
            if ws_template.max_row > 1:
                ws_template.delete_rows(2, ws_template.max_row)

            # 소스 데이터의 컬럼명과 템플릿의 컬럼명 매칭
            # 공백 제거 및 대소문자 무시로 매칭율 향상
            source_cols_clean = {str(col).strip(): col for col in df_source.columns}
            
            mapping = {} # template_col_idx -> source_col_name
            for t_idx, t_header in enumerate([cell.value for cell in ws_template[1]], start=1):
                if t_header is None: continue
                t_header_clean = str(t_header).strip()
                
                if t_header_clean in source_cols_clean:
                    mapping[t_idx] = source_cols_clean[t_header_clean]
                    # print(f"     - 컬럼 매칭: [{t_header}]")
                else:
                    print(f"     - 컬럼 없음: [{t_header}] (소스 파일에 해당 헤더가 없습니다)")

            # 데이터 복사
            for r_idx, row_data in enumerate(df_source.to_dict('records'), start=2):
                for t_col_idx, s_col_name in mapping.items():
                    val = row_data.get(s_col_name)
                    # NaN 처리
                    if pd.isna(val):
                        val = None
                    ws_template.cell(row=r_idx, column=t_col_idx, value=val)
        else:
            print(f"   - 건너뜀: '{sheet_name}' (소스 파일에 해당 시트가 없습니다)")

    print(f"4. 결과 저장 중: {output_file}")
    try:
        wb_template.save(output_file)
        print("🎉 모든 작업이 완료되었습니다!")
        print(f"생성된 파일: {output_file}")
    except Exception as e:
        print(f"Error: 저장 중 오류 발생: {e}")

if __name__ == "__main__":
    convert_to_template_robust()
